import requests
from bs4 import BeautifulSoup as bs
import openpyxl
import xls2xlsx
import logging

logging.basicConfig(filename="logs/calendar_logs.log", level=logging.INFO,
                    format=' %(asctime)s - %(levelname)s - %(message)s',
                    encoding="utf8")


def _format_xlsx(filename):
    wb = openpyxl.load_workbook(filename)
    for ws in wb:
        merged_cells = list(ws.merged_cells.ranges)
        for merged_cell_range in merged_cells:
            min_col, min_row, max_col, max_row = merged_cell_range.bounds
            merged_value = ws.cell(min_row, min_col).value
            ws.unmerge_cells(str(merged_cell_range))
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col,
                                    values_only=False):
                for cell in row:
                    cell.value = merged_value

    wb.save(filename)


def _xls_to_xlsx(name):
    xls_file = xls2xlsx.XLS2XLSX(name)
    xls_file.to_xlsx(name + "x")


def _format_link(link):
    if '//www.hse.ru' in link:
        link = link.replace('//www.hse.ru', 'http://students.perm.hse.ru')
    elif 'www.hse.ru' in link:
        link = link.replace('www.hse.ru', 'http://students.perm.hse.ru')
    elif 'http://students.perm.hse.ru' not in link:
        link = 'http://students.perm.hse.ru' + link

    return link


class Timetable(object):
    def __init__(self):
        self.num_of_tts = 0
        self.num_of_tts_eng = 0

    def update_timetable(self):
        logging.info("files update begin")
        tt_page = requests.get('http://students.perm.hse.ru/timetable/')
        soup = bs(tt_page.content, 'html.parser')
        links = []
        links_eng = []
        count = 0
        count_eng = 0

        for a in soup.find_all("a"):
            if "Расписание занятий (" in a.text:
                links.append(a.get("href"))
            elif "язык" in a.text:
                links_eng.append(a.get("href"))

        for link in links:
            link = _format_link(link)
            timetable = requests.get(link, allow_redirects=True)
            try:
                open(f"timetables/timetable{count}.xlsx", "wb").write(timetable.content)
                _format_xlsx(f"timetables/timetable{count}.xlsx")
            except:
                open(f"timetables/timetable{count}.xls", "wb").write(timetable.content)
                _xls_to_xlsx(f"timetables/timetable{count}.xls")
                _format_xlsx(f"timetables/timetable{count}.xlsx")
            count += 1

        for link in links_eng:
            link = _format_link(link)
            timetable = requests.get(link, allow_redirects=True)
            open(f"timetables/eng_timetable{count_eng}.xlsx", "wb").write(timetable.content)
            _format_xlsx(f"timetables/eng_timetable{count_eng}.xlsx")
            count_eng += 1

        self.num_of_tts = count
        logging.info(f"downloaded {self.num_of_tts} timetable files")
        self.num_of_tts_eng = count_eng
        logging.info(f"downloaded {self.num_of_tts_eng} english timetable files")

    def get_workbooks(self):
        for i in range(self.num_of_tts):
            yield openpyxl.load_workbook(f"timetables/timetable{i}.xlsx")

    def get_workbooks_eng(self):
        for i in range(self.num_of_tts_eng):
            yield openpyxl.load_workbook(f"timetables/eng_timetable{i}.xlsx")
